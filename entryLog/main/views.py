from django.shortcuts import render
from .models import *
from .forms import *
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from django.shortcuts import  get_object_or_404, redirect
from django.utils import timezone
from django.template.loader import get_template
from django.core.mail import EmailMessage
import openpyxl as xl
import datetime, random

def home(request):
    
    return render(request, 'home.html')

def entry_email(visitor_id):    
    
    visitor = Entry.objects.get(pk=visitor_id)
    try :
        template = get_template('guest_arrival.txt')  # Body of email is generated there                  
        context = {
            'visitor' : visitor,
        }
        content = template.render(context) 
        email = EmailMessage(
            "Guest Arrival",
            content,
            "IIT" +'',
            [ visitor.host_name.email ],
            headers = {'Reply-To': 'd21bhardwaj@gmail.com' }
        )
        email.send()
        otp = visitor.verification
        message = str("Guest has been successfully checked in. Please note your 4 digit OTP for checkout | "+otp+" |")

    except Exception as e :
        print(type(e)) # Error can be checked at console. #
    #   If email is not send entered will not be created.
    #   visitor.delete() --- #
        otp = visitor.verification
        message = str("There has been an error in sending email,But you have been checked in Please note your 4 digit OTP for checkout | "+otp+" |")
        
    return message

def guest_entry(request):
    
    form_class = EntryForm
    form = form_class(request.POST or None)
    if request.method == 'POST':          
        if form.is_valid():            
            
            # try is used to check if there exsist an entered which has not  been checked out
            try:
                visitor_email = form.cleaned_data.get('visitor_email')            
                guest = Entry.objects.get(visitor_email = visitor_email, left_office = False)                
                return HttpResponse("Guest has not checked out of the previous visit. Please check out before a new check in.")       
            except ObjectDoesNotExist:
                visitor= form.save(commit=False)
            #   Generating otp to be used while checkout
                otp = random.randint(1000,9999)
                visitor.verification = otp
                visitor = form.save()
            #   To check if the email has been successfully sent
                message = Entry_email(visitor.id)
                return HttpResponse(message)   #---- R1 ----#
    #---- If you want to check it without email functioning  please comment R1 and Un comment the below return ---#
            #    return HttpResponse("Guest has been successfully checked in. Please note your 4 digit OTP for checkout | "+otp+" |")
        
        else:
            print(form.errors)        
        
    return render(request, 'form.html', {
        'form': form,'header':'Guest Entry'
    })

def exit_email(visitor_id):
    
    visitor = Entry.objects.get(pk=visitor_id)
    try :
        template = get_template('guest_departure.txt')                
        context = {
            'visitor' : visitor,
            'address' : 'IIT Guwahati, Guwahati'
        }
        content = template.render(context)      # Body of email is generated there # 
        email = EmailMessage(
            "Guest Arrival",
            content,
            "IIT" +'',
            [ visitor.visitor_email ],
            headers = {'Reply-To': 'd21bhardwaj@gmail.com' }
        )
        email.send()
        message = str("Thanks for visiting. An email has been succesfully generated regarding your visit.")
        
    except Exception as e :
        print(type(e))
        message = str("We couldnot send you the email, But you have been succesfully checked out at "
            + str(visitor.out_time) + " Thanks for visiting us.")
        
    return message

def guest_check_out(request, guest_id):
    form_class = ExitForm
    form = form_class(request.POST or None)
    guest = Entry.objects.filter(left_office= False )
    visitor = get_object_or_404(Entry, pk=guest_id)
    if visitor.left_office == True :
        return HttpResponse("Guest has already checked out.")
    elif request.method == 'POST': 
        if form.is_valid():
            
            otp = form.cleaned_data['verification']
            if visitor.verification == otp :  
                visitor.out_time = datetime.datetime.now()
                visitor.left_office = True 
                visitor.save() 
                message = exit_email(visitor.id)
            else :
                message = str("Please enter the correct OTP")
            return HttpResponse(message)
        else:
            print(form.errors)
    return render(request, 'guest_log.html', {'guest' : guest,'guest_id':guest_id, 'form':form})

def all_guest(request):
    guest = Entry.objects.filter(left_office= False )
    return render(request, 'guest_log.html', {'guest' : guest, })

def all_employee(request):
    
    employees = Employee.objects.all()
    return render(request, 'employee_log.html', {'employees' : employees})

def employee_availablity(request, employee_id):
    
    employee = get_object_or_404(Employee, pk=employee_id)
    if(employee.in_office):
        employee.in_office = False
    else :
        employee.in_office = True
    employee.save()
    return redirect(all_employee)

def employee_upload(request):
    
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = xl.load_workbook(excel_file)
        sheet = wb.active
        # print(sheet)
        max_r = sheet.max_row
        dic = []
         
        for i in range(1, max_r+1):
            if Employee.objects.filter(email=str(sheet.cell(row=i,column= 5).value)).first() :
                email=str(sheet.cell(row=i,column= 5).value)
                error = str("Employee with email :"+ email +" exist already. Edit row "+ str(i)+" before continuing")
                return HttpResponse(error)
        
        for i in range(1,max_r+1): 
            if(i!=1):                
                dic.append(Employee(
                title         = str(sheet.cell(row=i,column=1).value),      
                name          = str(sheet.cell(row=i,column=2).value),
                designation   = str(sheet.cell(row=i,column=3).value),  
                phone_no      = str(sheet.cell(row=i,column=4).value), 
                email         = str(sheet.cell(row=i,column=5).value), 
                department    = str(sheet.cell(row=i,column=6).value), 
                room_no       = str(sheet.cell(row=i,column=7).value),
                ))        
        Employee.objects.bulk_create(dic)
        return HttpResponse("Data created at server for"+" ")

    return render(request,"employee_data.html")

def log(request):

    guest = Entry.objects.all()
    return render(request, 'log.html', {'guest':guest})


